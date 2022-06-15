import openpyxl as xl

wb = xl.load_workbook("Määritysrajat.xlsx")
ws = wb.active

limits_sulate = {}
limits_puriste = {}
for row in ws.iter_rows(min_row=4, min_col=2, max_col=4):
    if not row[0].value:
        break
    limits_sulate[row[0].value] = (row[1].value, row[2].value)

for row in ws.iter_rows(min_row=4, min_col=6, max_col=8):
    if not row[0].value:
        break
    limits_puriste[row[0].value] = (row[1].value, row[2].value)
    
print(limits_puriste)