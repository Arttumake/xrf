import csv
import os
import glob
import shutil
import datetime
import ctypes
import warnings

import openpyxl as xl
from openpyxl.styles.fonts import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
"""
Done by: Arttu Mäkelä 
For: CRS Laboratories
Started: June 2022

This script reads csv file in folder and outputs an Excel report based on
the excel template file. It renames the input CSV-file, copies the Excel report to
a subdirectory and moves the CSV-file to another. 
"""


offset: int  # how many rows in template before csv compounds, assigned once template is determined
substance_row = 11  # the row where all the compounds are listed in excel templates

# templates used in this script
uniquant_template = "uniquant.xlsx"
puriste_template = "puriste.xlsx"
sulate_template = "sulate.xlsx"
puriste_sulate_template = "puriste_sulate.xlsx"

määritys_rajat_xl = "määritysrajat.xlsx"  # name of the määritysrajat-excel (only needed for puriste/sulate)
csv_dir = "CSV"  # CVS directory name
excel_dir = "Raportit"  # Excel report directory name

# group up all the csv files in this directory to a list
parent_path = os.path.abspath("..")
csv_files = glob.glob(os.path.join(parent_path, "*.csv"))

# show error window if csv-file not in correct folder
if not csv_files:
    ctypes.windll.user32.MessageBoxW(
        0, "Place CSV-file Export-folder for script to run", "Error", 0
    )

excels = []
csvs = []
date_format = "%Y-%b-%d %X"  # how CSV-file represents a date


for num, file in enumerate(csv_files):
    with open(file) as csv_file:
        file_reader = csv.reader(csv_file, delimiter=",")
        dates_rows = {}
        names = []
        # get each date in csv and put them in dictionary as datetime object keys
        # and row numbers being the values.
        # determine the correct excel template to load
        for row_num, row in enumerate(file_reader):
            if row_num == 0:
                if "Oxides" in row[0] or "Sulphides" in row[0]:
                    template = uniquant_template
                    offset = 6
                elif "PhosphateConcentrateMajors" in row[0]:
                    template = sulate_template
                    offset = 2
                elif "PhosphateRocks_PP" in row[0]:
                    template = puriste_template
                    offset = 2

            date = datetime.datetime.strptime(row[3], date_format)
            dates_rows[date] = row_num + 1
            names.append(row[1])

        # check if batch name appears twice and if so, swap template to puriste_sulate
        duplicates = set([name for name in names if names.count(name) == 2])
        if len(duplicates) == len(dates_rows) / 2:
            template = puriste_sulate_template

        # load the template and set active worksheet
        wb = xl.load_workbook(template)
        ws = wb.active

        # dictionary to hold compound as key and its column number as value
        compound_order = {}

        # read excel-template and get all the compounds in it to a dict
        for rows in ws.iter_rows(
            min_row=substance_row, max_row=substance_row, min_col=offset + 1
        ):
            for col_idx, cell in enumerate(rows):
                compound_order[cell.value] = col_idx + offset + 1

        # check if template excel is puriste/sulate and assign limits to compounds
        if template != uniquant_template:
            wb_limits = xl.load_workbook(määritys_rajat_xl)
            ws_limits = wb_limits.active
            if template == sulate_template or template == puriste_sulate_template:
                limits_sulate = {}
                for rows in ws_limits.iter_rows(min_row=4, min_col=2, max_col=4):
                    if not rows[0].value:
                        break
                    limits_sulate[rows[0].value] = (rows[1].value, rows[2].value)
            if template == puriste_template or template == puriste_sulate_template:
                limits_puriste = {}
                for rows in ws_limits.iter_rows(min_row=4, min_col=6, max_col=8):
                    if not rows[0].value:
                        break
                    limits_puriste[rows[0].value] = (rows[1].value, rows[2].value)

        # remove trailing none-key from dict if it exists
        compound_order.pop(None, None)

        # construct the date-portion of the report file's name
        current = datetime.datetime.now()
        file_date = f"{current.day}.{current.month}.{current.year}"
        # more exact time in the case that multiple reports are created from one csv
        file_date_exact = (
            file_date + f" {current.hour} {current.minute} {current.second}"
        )

        dates = []
        # sort the dates
        for key, value in dates_rows.items():
            dates.append(key)
        sorted_dates_rows = {}
        dates.sort()

        for i, key in enumerate(dates_rows.keys()):
            sorted_dates_rows[key] = dates.index(key) + 1

        # row_order shows where each row in csv is placed in excel report
        row_order = [item for item in sorted_dates_rows.values()]
        csv_file.seek(0)  # reset iterator to beginning

        # keep track of sample names and their row for puriste_sulate-template
        unique_samples = {}

        # booleans to determine how method cell in report should be named
        oxides_used = False
        sulphides_used = False
        # list to hold all unique methods used in a uniquant csv
        uquant_methods = []  
        # loop through samples (sorted by date) and csv iterable
        for index, (row_num, row) in enumerate(zip(row_order, file_reader)):
            for col, value in enumerate(row):  # loop through each value in csv row
                # +2 for the extra 2 rows under compounds
                current_row = substance_row + row_num + 2
                # place all unique methods in uniquant csv to list, place them in report later
                if col == 0 and value not in uquant_methods:
                    uquant_methods.append(value)

                # check the cells containing values in csv and place them in report
                elif col > 3 and col % 2 != 0:
                    try:
                        # set column number based on compound's order in excel template
                        this_column = compound_order[row[col - 1]]
                        if template != puriste_sulate_template:
                            # insert value from csv to correct row/column in excel report
                            ws.cell(row=current_row, column=this_column).value = float(
                                value
                            )

                            # place values to Fe and Mn columns instead of Fe2O3 or MnO if sulphides method used in csv row
                            if (
                                template == uniquant_template
                                and "sulphides" in row[0].lower()
                            ):
                                if row[col - 1] == "Fe2O3":
                                    fe = compound_order["Fe"]
                                    ws.cell(row=current_row, column=fe).value = float(
                                        value
                                    )
                                    ws.cell(
                                        row=current_row, column=this_column
                                    ).value = None
                                elif row[col - 1] == "MnO":
                                    mn = compound_order["Mn"]
                                    ws.cell(row=current_row, column=mn).value = float(
                                        value
                                    )
                                    ws.cell(
                                        row=current_row, column=this_column
                                    ).value = None

                            if (
                                template == uniquant_template
                                and "oxides" in row[0].lower()
                            ):
                                if row[col - 1] == "Fe2O3":
                                    fe = compound_order["Fe"]
                                    ws.cell(row=current_row, column=fe).value = round(
                                        0.69945 * float(value), 3
                                    )

                        elif template == puriste_sulate_template:
                            sample_name = row[1]  # refers to sample name in csv
                            sub_method = ws.cell(
                                row=13, column=this_column
                            ).value.strip("\xa0")
                            if sample_name not in unique_samples.keys():
                                unique_samples[sample_name] = current_row
                            # place value to report given the condition and if its the 2nd time the
                            # sample name appears, place its value in that sample name's row instead
                            if (
                                "LBF" in sub_method[:3]
                                and "PhosphateConcentrateMajors" in row[0]
                                or "PP" in sub_method[:2]
                                and "PhosphateRocks" in row[0]
                            ):
                                if sample_name not in unique_samples.keys():
                                    ws.cell(
                                        row=current_row, column=this_column
                                    ).value = float(value)
                                else:
                                    ws.cell(
                                        row=unique_samples[sample_name],
                                        column=this_column,
                                    ).value = float(value)

                    # catch all compounds not defined in the dictionary and place
                    # their values after "Sum Before Norm." cell
                    except KeyError:
                        this_column = len(compound_order) + 2
                        compound_cell = ws.cell(row=substance_row, column=this_column)
                        compound = row[col - 1]

                        def place_value(this_column):
                            try:
                                ws.cell(
                                    row=current_row, column=this_column
                                ).value = float(value)
                                compound_cell.font = Font(
                                    name="Arial", bold=True, size=10
                                )
                                compound_cell.alignment = Alignment(horizontal="center")
                                compound_cell.border = Border(bottom=Side(style="thin"))
                                # styling for the 2 rows below compound
                                below_compound = ws.cell(
                                    row=substance_row + 1, column=this_column
                                )
                                below_compound.value = "(%)"
                                below_compound.alignment = Alignment(
                                    horizontal="center"
                                )
                                below_compound.font = Font(name="Arial", size=8)
                                pp_xrf12 = ws.cell(
                                    row=substance_row + 2, column=this_column
                                )
                                pp_xrf12.alignment = Alignment(horizontal="center")
                                pp_xrf12.value = ws.cell(
                                    row=substance_row + 2, column=2
                                ).value
                                pp_xrf12.font = Font(name="Arial", size=8)

                            except ValueError:
                                return

                        # check if compound already placed on column
                        if not compound_cell.value:
                            compound_cell.value = compound
                            place_value(this_column)
                        elif compound_cell.value == compound:
                            place_value(this_column)
                        # if compound is a different compound, put its value on the next empty or matching column
                        elif compound_cell.value != compound:
                            is_avail = False
                            col_idx = 1
                            while not is_avail:
                                compound_cell = ws.cell(
                                    row=substance_row, column=this_column + col_idx
                                )
                                debug = compound_cell.value
                                if (
                                    compound == compound_cell.value
                                    or not compound_cell.value
                                ):
                                    compound_cell.value = compound
                                    place_value(this_column + col_idx)
                                    is_avail = True
                                col_idx += 1

                # sample name column from csv to excel report
                elif col == 1:
                    ws.cell(row=current_row, column=1).value = value
                    ws.cell(row=current_row, column=1).alignment = Alignment(
                        horizontal="right"
                    )

                # get SID2 from first row, column 3
                elif col == 2 and index == 0:
                    sid2 = value  # to be used in naming the report
                    ws.cell(
                        row=4, column=2
                    ).value = value  # add sid2 to excel report as batch name
                    
    # place all methods in uniquant csv to the correct field in excel report
    if template == uniquant_template:
        ws.cell(row=5, column=2).value = ", ".join(uquant_methods)
    
    # check limits for sulate values and overwrite if over/under
    if template == sulate_template:
        for key, value in limits_sulate.items():
            for col in ws.iter_cols(
                min_row=substance_row + 3,
                min_col=compound_order[key],
                max_col=compound_order[key],
            ):
                for cell in col:
                    if cell.value or cell.value == 0:
                        try:
                            if cell.value < limits_sulate[key][0]:
                                cell.value = f"< {limits_sulate[key][0]}"
                            elif cell.value > limits_sulate[key][1]:
                                cell.value = f"*{limits_sulate[key][1]}"
                        except TypeError:
                            continue

    # check limits for puriste values and overwrite if over/under
    elif template == puriste_template:
        for key, value in limits_puriste.items():
            for col in ws.iter_cols(
                min_row=substance_row + 3,
                min_col=compound_order[key],
                max_col=compound_order[key],
            ):
                for cell in col:
                    if cell.value or cell.value == 0:
                        try:
                            if cell.value < limits_puriste[key][0]:
                                cell.value = f"< {limits_puriste[key][0]}"
                            elif cell.value > limits_puriste[key][1]:
                                cell.value = f"> {limits_puriste[key][1]}"
                        except TypeError:
                            continue
                            
    # check for None-values in excel report and insert value to them
    if template != puriste_sulate_template:
        if template == uniquant_template:
            exclude = [
                "Mn",
                "MnO",
                "Fe",
                "Fe2O3",
            ]  # don't put a value in cells under these compounds
            exclude_cols = [
                compound_order[comp] for comp in exclude
            ]  # exclude list converted to column number list
            for row in ws.iter_rows(
                min_row=substance_row + 3,
                min_col=offset + 1,
                max_col=len(compound_order) + offset,
            ):
                if row[0].value:  # check that row has a sample name
                    for cell in row:
                        if not cell.value and cell.column not in exclude_cols:
                            cell.value = "< 0.001"
                        cell.alignment = Alignment(horizontal="right")
        else:
            for row in ws.iter_rows(
                min_row=substance_row + 3,
                min_col=offset + 1,
                max_col=len(compound_order) + offset,
            ):
                if row[0].value:  # check that row has a sample name
                    for cell in row:
                        if not cell.value:
                            cell.value = "< 0.001"
                        cell.alignment = Alignment(horizontal="right")

    # delete duplicate rows with no values from excel report if puriste_sulate template
    else:
        for col in ws.iter_cols(
            min_row=substance_row + 3,
            max_row=substance_row + 3 + len(row_order),
            min_col=offset + 1,
            max_col=offset + 1,
        ):
            for cell in col:
                if not cell.value:
                    ws.delete_rows(cell.row)

    # check limits for puriste_sulate template
    if template == puriste_sulate_template:
        starting_col = 2
        for col in ws.iter_cols(
            min_row=substance_row + 3, min_col=2, max_col=len(compound_order) + 2
        ):
            unstripped_method = ws.cell(
                row=substance_row + 2, column=starting_col
            ).value

            if unstripped_method:
                filter_method = unstripped_method.strip("\xa0")
                compound = ws.cell(row=substance_row, column=starting_col).value

                for cell in col:
                    cell.alignment = Alignment(horizontal="right")
                    if cell.value or cell.value == 0:
                        if "PP" in filter_method[:2]:
                            try:
                                if cell.value < limits_puriste[compound][0]:
                                    cell.value = f"< {limits_puriste[compound][0]}"
                            except TypeError:
                                continue
                            try:
                                if cell.value > limits_puriste[compound][1]:
                                    cell.value = f"> {limits_puriste[compound][1]}"
                            except TypeError:
                                continue
                        elif "LBF" in filter_method[:3]:
                            try:
                                if cell.value < limits_sulate[compound][0]:
                                    cell.value = f"< {limits_sulate[compound][0]}"
                            except TypeError:
                                continue
                            try:
                                if cell.value > limits_sulate[compound][1]:
                                    cell.value = f"*{cell.value}"
                            except TypeError:
                                continue
            starting_col += 1

    forbidden = "/\?:|<>"  # characters that can't be used in file name
    sid2_clean = "".join(
        " " if letter in forbidden else letter for letter in sid2
    ).rstrip()

    # rename csv file and save a new excel file

    excel_name = f"{sid2_clean} - {file_date}.xlsx"
    csv_name = f"{sid2_clean} - {file_date}.csv"

    csv_path = os.path.join(parent_path, csv_name)
    os.rename(file, csv_path)
    wb.save(excel_name)
    wb.close()

    excel_path = os.path.join(os.getcwd(), excel_name)
    csvs.append(csv_path)
    excels.append(excel_path)


def move_files(names: list, dest: str, cwd=os.getcwd(), file_type=".csv"):
    """Moves a list of files to the destination folder, script should be
    in the same folder as the files being moved.
    Renames copied file if file already exist in destination folder by
    adding more accurate timestamp

    args:
        - names: list of file names in path format
        - dest: destination folder (not path, string only)
        - cwd: location of the file that is to be moved (in relation to this script)
        - file_type: what type of files are moved, in case of duplicates
        and they need renaming

    """
    for name in names:
        dir = os.path.join(os.path.abspath(".."), dest)
        tail = os.path.split(name)
        file_path = os.path.join(dir, tail[1])
        if not os.path.exists(dir):
            os.mkdir(dir)
        if os.path.exists(file_path):
            file_path = os.path.join(cwd, name)
            new_name = f"{sid2_clean} - {file_date_exact}{file_type}"
            os.rename(file_path, os.path.join(cwd, new_name))
            shutil.move(os.path.join(cwd, new_name), dir)
        else:
            shutil.move(name, dir)


#move_files(excels, excel_dir, file_type=".xlsx")
#move_files(csvs, csv_dir, cwd=parent_path)
